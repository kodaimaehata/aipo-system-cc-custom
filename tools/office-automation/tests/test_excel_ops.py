from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from office_automation.excel_ops import edit, read



def _build_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet1["A1"] = "header"
    sheet1["B2"] = 42
    sheet1.merge_cells("C1:D1")
    sheet1["C1"] = "merged"

    sheet2 = workbook.create_sheet(title="Data")
    sheet2["A1"] = "name"
    sheet2["A2"] = "alice"
    sheet2["B2"] = "engineer"

    workbook.save(path)
    workbook.close()
    return path



def test_read_returns_structured_workbook_summary_for_xlsx(tmp_path: Path) -> None:
    workbook_path = _build_workbook(tmp_path / "report.xlsx")

    result = read(workbook_path)

    assert result["file_path"] == str(workbook_path)
    assert result["extension"] == "xlsx"
    assert result["sheet_count"] == 2
    assert result["sheet_names"] == ["Sheet1", "Data"]

    first_sheet = result["sheets"][0]
    assert first_sheet["name"] == "Sheet1"
    assert first_sheet["max_row"] == 2
    assert first_sheet["max_column"] == 4
    assert first_sheet["merged_ranges"] == ["C1:D1"]
    assert first_sheet["populated_cells"] == [
        {"cell": "A1", "row": 1, "column": 1, "value": "header"},
        {"cell": "C1", "row": 1, "column": 3, "value": "merged"},
        {"cell": "B2", "row": 2, "column": 2, "value": 42},
    ]

    second_sheet = result["sheets"][1]
    assert second_sheet["name"] == "Data"
    assert second_sheet["populated_cells"] == [
        {"cell": "A1", "row": 1, "column": 1, "value": "name"},
        {"cell": "A2", "row": 2, "column": 1, "value": "alice"},
        {"cell": "B2", "row": 2, "column": 2, "value": "engineer"},
    ]



def test_edit_applies_cell_and_sheet_operations_to_authoritative_output_path(tmp_path: Path) -> None:
    source = _build_workbook(tmp_path / "report.xlsx")
    output = tmp_path / "edited" / "report.xlsx"

    saved_path = edit(
        source,
        {
            "output_path": str(output),
            "operations": [
                {"op": "set_cell", "sheet": "Sheet1", "cell": "A2", "value": "updated"},
                {"op": "clear_cell", "sheet": "Sheet1", "cell": "B2"},
                {"op": "rename_sheet", "from": "Data", "to": "Summary"},
                {"op": "add_sheet", "name": "Notes", "index": 1},
                {"op": "delete_sheet", "name": "Summary"},
            ],
            "output_dir": str(tmp_path / "ignored-metadata"),
        },
    )

    assert saved_path == output
    assert output.exists()
    assert source.exists()

    original = load_workbook(source)
    try:
        assert original.sheetnames == ["Sheet1", "Data"]
        assert original["Sheet1"]["A2"].value is None
        assert original["Sheet1"]["B2"].value == 42
    finally:
        original.close()

    edited = load_workbook(output)
    try:
        assert edited.sheetnames == ["Sheet1", "Notes"]
        assert edited["Sheet1"]["A2"].value == "updated"
        assert edited["Sheet1"]["B2"].value is None
    finally:
        edited.close()



def test_edit_preserves_original_by_copying_before_edit(tmp_path: Path) -> None:
    source = _build_workbook(tmp_path / "original.xlsx")
    output = tmp_path / "copies" / "custom-name.xlsx"

    saved_path = edit(
        source,
        {
            "output_path": output,
            "copy_before_edit": True,
            "operations": [
                {"op": "set_cell", "sheet": "Sheet1", "cell": "B2", "value": "new value"},
            ],
        },
    )

    assert saved_path == output
    assert source.read_bytes() != output.read_bytes()

    original = load_workbook(source)
    try:
        assert original["Sheet1"]["B2"].value == 42
    finally:
        original.close()

    edited = load_workbook(output)
    try:
        assert edited["Sheet1"]["B2"].value == "new value"
    finally:
        edited.close()



def test_edit_does_not_leave_output_artifact_when_copy_before_edit_operation_fails(tmp_path: Path) -> None:
    source = _build_workbook(tmp_path / "original.xlsx")
    output = tmp_path / "copies" / "failed.xlsx"

    with pytest.raises(ValueError, match=r"Excel operation #1 references missing sheet 'Missing'\."):
        edit(
            source,
            {
                "output_path": output,
                "copy_before_edit": True,
                "operations": [
                    {"op": "set_cell", "sheet": "Missing", "cell": "A1", "value": "new value"},
                ],
            },
        )

    assert not output.exists()
    assert source.exists()

    original = load_workbook(source)
    try:
        assert original["Sheet1"]["A1"].value == "header"
    finally:
        original.close()



@pytest.mark.parametrize("extension", ["xls", "xlsb"])
def test_read_rejects_legacy_excel_extensions(tmp_path: Path, extension: str) -> None:
    legacy_path = tmp_path / f"legacy.{extension}"
    legacy_path.write_text("legacy")

    with pytest.raises(
        ValueError,
        match=rf"Excel V1 scope does not support '\.{extension}' files\. Supported Excel extensions: xlsm, xlsx\.",
    ):
        read(legacy_path)



@pytest.mark.parametrize(
    ("instructions", "exception_type", "message"),
    [
        (
            {"operations": []},
            ValueError,
            r"Excel edit instructions must include 'output_path'; runtime save-path fallback is not supported\.",
        ),
        (
            {"output_path": "out.xlsx", "operations": "not-a-list"},
            TypeError,
            r"Excel edit instructions field 'operations' must be a list\.",
        ),
        (
            {"output_path": "out.xlsx", "operations": [{"op": "rename_sheet", "from": "Sheet1"}]},
            ValueError,
            r"Excel operation #1 field 'to' must be a non-empty string\.",
        ),
        (
            {"output_path": "out.xlsx", "operations": [{"op": "set_cell", "sheet": "Sheet1", "cell": "12", "value": "x"}]},
            ValueError,
            r"Excel operation #1 has invalid cell reference '12'\.",
        ),
        (
            {"output_path": "out.xlsx", "operations": [{"op": "set_cell", "sheet": "Sheet1", "cell": "A1", "value": ["x"]}]},
            TypeError,
            r"Excel operation #1 field 'value' is invalid: Excel set_cell values must be one of: string, int, float, bool, Decimal, date, datetime, time, timedelta, or None; got list\.",
        ),
        (
            {"output_path": "out.xlsx", "operations": [{"op": "run_macro", "name": "Cleanup"}]},
            ValueError,
            r"Excel operation #1 requests unsupported VBA or macro editing scope: 'run_macro'\.",
        ),
        (
            {
                "output_path": "out.xlsx",
                "operations": [
                    {
                        "op": "set_cell",
                        "sheet": "Sheet1",
                        "cell": "A1",
                        "value": "x",
                        "vba_module": "Module1",
                    }
                ],
            },
            ValueError,
            r"Excel operation #1 request unsupported VBA or macro editing scope via 'vba_module'\.",
        ),
        (
            {"output_path": "out.xlsx", "operations": [], "options": {"edit_vba": True}},
            ValueError,
            r"Excel edit options request unsupported VBA or macro editing scope via 'edit_vba'\.",
        ),
    ],
)
def test_edit_rejects_malformed_or_unsupported_instruction_payloads(
    tmp_path: Path,
    instructions: dict,
    exception_type: type[Exception],
    message: str,
) -> None:
    source = _build_workbook(tmp_path / "report.xlsx")

    with pytest.raises(exception_type, match=message):
        edit(source, instructions)



def test_edit_rejects_same_output_path_when_copy_before_edit_is_enabled(tmp_path: Path) -> None:
    source = _build_workbook(tmp_path / "report.xlsx")

    with pytest.raises(
        ValueError,
        match=r"Excel edit requested copy-before-edit preservation, but output_path points to the source file\.",
    ):
        edit(
            source,
            {
                "output_path": str(source),
                "copy_before_edit": True,
                "operations": [{"op": "set_cell", "sheet": "Sheet1", "cell": "A1", "value": "new"}],
            },
        )



def test_xlsm_read_and_edit_smoke_flow_uses_supported_extension_path(tmp_path: Path) -> None:
    source = _build_workbook(tmp_path / "macrobook.xlsm")
    output = tmp_path / "out" / "macrobook.xlsm"

    read_result = read(source)
    assert read_result["extension"] == "xlsm"
    assert read_result["sheet_names"] == ["Sheet1", "Data"]

    saved_path = edit(
        source,
        {
            "output_path": output,
            "operations": [
                {"op": "set_cell", "sheet": "Sheet1", "cell": "A1", "value": "macro-safe"},
            ],
        },
    )

    assert saved_path == output

    edited = load_workbook(output, keep_vba=True)
    try:
        assert edited["Sheet1"]["A1"].value == "macro-safe"
        assert edited.sheetnames == ["Sheet1", "Data"]
    finally:
        if getattr(edited, "vba_archive", None) is not None:
            edited.vba_archive.close()
        edited.close()
