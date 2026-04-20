"""Helpers for xlsx and xlsm workbook operations."""

from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime, time, timedelta
from decimal import Decimal
import math
from os import PathLike
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["read", "edit"]

_SUPPORTED_EXTENSIONS = frozenset({"xlsx", "xlsm"})
_UNSUPPORTED_LEGACY_EXTENSIONS = frozenset({"xls", "xlsb"})
_SUPPORTED_EXTENSIONS_TEXT = ", ".join(sorted(_SUPPORTED_EXTENSIONS))
_SUPPORTED_OPERATIONS = frozenset(
    {"set_cell", "clear_cell", "rename_sheet", "add_sheet", "delete_sheet"}
)
_UNSUPPORTED_MACRO_KEYS = frozenset(
    {
        "macro",
        "macros",
        "macro_name",
        "macro_names",
        "macro_project",
        "vba",
        "vba_module",
        "vba_modules",
        "vba_project",
        "edit_macro",
        "edit_macros",
        "edit_vba",
        "edit_vba_project",
    }
)


def read(file_path: str | Path) -> dict:
    """Read an xlsx/xlsm workbook and return structured data."""
    source = _validate_source_path(file_path)
    workbook = _load_excel_workbook(source)
    try:
        worksheets = list(workbook.worksheets)
        return {
            "file_path": str(source),
            "extension": _path_extension(source),
            "sheet_count": len(worksheets),
            "sheet_names": [sheet.title for sheet in worksheets],
            "sheets": [_serialize_sheet(sheet) for sheet in worksheets],
        }
    finally:
        _close_workbook(workbook)



def edit(file_path: str | Path, instructions: dict) -> Path:
    """Apply xlsx/xlsm workbook edits and return the saved output path."""
    source = _validate_source_path(file_path)
    payload = _validate_instruction_envelope(instructions)
    output_path = _validate_output_path(payload["output_path"], source)
    copy_before_edit = payload["copy_before_edit"]

    save_path = _prepare_edit_target(source, output_path, copy_before_edit=copy_before_edit)
    workbook = _load_excel_workbook(source)

    try:
        for index, operation in enumerate(payload["operations"], start=1):
            _apply_operation(workbook, operation, index=index)
        _save_workbook(workbook, source=source, save_path=save_path)
    finally:
        _close_workbook(workbook)

    return output_path



def _coerce_path(value: str | PathLike[str]) -> Path:
    return Path(value)



def _validate_source_path(file_path: str | PathLike[str]) -> Path:
    path = _coerce_path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file '{path}' does not exist.")
    if path.is_symlink() or not path.is_file():
        raise ValueError(f"Excel file '{path}' is not a regular file.")

    extension = _path_extension(path)
    if extension in _UNSUPPORTED_LEGACY_EXTENSIONS:
        raise ValueError(
            f"Excel V1 scope does not support '.{extension}' files. "
            "Supported Excel extensions: xlsm, xlsx."
        )
    if extension not in _SUPPORTED_EXTENSIONS:
        display_extension = extension or "<none>"
        raise ValueError(
            f"Excel file '{path}' has unsupported extension '{display_extension}'. "
            f"Supported Excel extensions: {_SUPPORTED_EXTENSIONS_TEXT}."
        )
    return path



def _validate_instruction_envelope(instructions: dict) -> dict:
    if not isinstance(instructions, dict):
        raise TypeError("Excel edit instructions must be a dictionary.")

    _reject_macro_scope(instructions)

    operations = instructions.get("operations")
    if operations is None:
        raise ValueError("Excel edit instructions must include an 'operations' list.")
    if not isinstance(operations, list):
        raise TypeError("Excel edit instructions field 'operations' must be a list.")

    output_path = instructions.get("output_path")
    if output_path in (None, ""):
        raise ValueError(
            "Excel edit instructions must include 'output_path'; runtime save-path fallback is not supported."
        )

    copy_before_edit = instructions.get("copy_before_edit", True)
    if not isinstance(copy_before_edit, bool):
        raise TypeError("Excel edit instructions field 'copy_before_edit' must be a boolean.")

    options = instructions.get("options", {})
    if not isinstance(options, dict):
        raise TypeError("Excel edit instructions field 'options' must be a dictionary.")
    _reject_macro_scope(options, context="Excel edit options")

    normalized_operations = []
    for index, operation in enumerate(operations, start=1):
        normalized_operations.append(_validate_operation(operation, index=index))

    return {
        "operations": normalized_operations,
        "output_path": output_path,
        "copy_before_edit": copy_before_edit,
        "options": options,
    }



def _validate_output_path(output_path: str | PathLike[str], source: Path) -> Path:
    path = _coerce_path(output_path)
    if not path.name:
        raise ValueError("Excel output_path must identify a file path.")

    extension = _path_extension(path)
    if extension != _path_extension(source):
        raise ValueError(
            "Excel output_path extension must match the source workbook extension "
            f"('.{_path_extension(source)}')."
        )
    return path



def _prepare_edit_target(source: Path, output_path: Path, *, copy_before_edit: bool) -> Path:
    source_resolved = source.resolve()
    output_resolved = output_path.resolve(strict=False)
    same_location = source_resolved == output_resolved

    if copy_before_edit:
        if same_location:
            raise ValueError(
                "Excel edit requested copy-before-edit preservation, but output_path points to the source file."
            )
        if output_path.exists():
            raise FileExistsError(f"Excel output path '{output_path}' already exists.")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        return output_path

    if output_path.exists() and not same_location:
        raise FileExistsError(f"Excel output path '{output_path}' already exists.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path



def _load_excel_workbook(path: Path) -> Workbook:
    try:
        return load_workbook(path, keep_vba=_path_extension(path) == "xlsm")
    except Exception as exc:  # pragma: no cover - library-specific failures vary by file.
        raise ValueError(f"Failed to load Excel workbook '{path}': {exc}") from exc



def _close_workbook(workbook: Workbook) -> None:
    vba_archive = getattr(workbook, "vba_archive", None)
    if vba_archive is not None:
        vba_archive.close()
    workbook.close()



def _serialize_sheet(sheet: Worksheet) -> dict:
    return {
        "name": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "merged_ranges": [str(cell_range) for cell_range in sheet.merged_cells.ranges],
        "populated_cells": _collect_populated_cells(sheet),
    }



def _collect_populated_cells(sheet: Worksheet) -> list[dict]:
    cell_map = getattr(sheet, "_cells", {})
    populated_cells: list[dict] = []
    for cell in sorted(cell_map.values(), key=lambda item: (item.row, item.column)):
        if cell.value is None:
            continue
        populated_cells.append(
            {
                "cell": cell.coordinate,
                "row": cell.row,
                "column": cell.column,
                "value": _normalize_cell_value(cell.value),
            }
        )
    return populated_cells



def _normalize_cell_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value



def _is_supported_cell_value(value) -> bool:
    return _cell_value_validation_error(value) is None



def _cell_value_validation_error(value) -> TypeError | ValueError | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, str):
        if ILLEGAL_CHARACTERS_RE.search(value):
            return ValueError(
                "Excel set_cell values cannot contain illegal XML control characters."
            )
        return None
    if isinstance(value, float):
        if not math.isfinite(value):
            return ValueError("Excel set_cell numeric values must be finite.")
        return None
    if isinstance(value, (int, Decimal, datetime, date, time, timedelta)):
        return None
    return TypeError(
        "Excel set_cell values must be one of: string, int, float, bool, Decimal, "
        f"date, datetime, time, timedelta, or None; got {type(value).__name__}."
    )



def _reject_macro_scope(payload: dict, *, context: str = "Excel edit instructions") -> None:
    if not isinstance(payload, dict):
        return

    for key, value in payload.items():
        normalized_key = str(key).strip().lower().replace("-", "_")
        if normalized_key in _UNSUPPORTED_MACRO_KEYS or "macro" in normalized_key or "vba" in normalized_key:
            if _macro_signal_present(value):
                raise ValueError(
                    f"{context} request unsupported VBA or macro editing scope via '{key}'."
                )



def _macro_signal_present(value) -> bool:
    if value is None or value is False:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, Iterable) and not isinstance(value, (bytes, bytearray, dict)):
        return any(_macro_signal_present(item) for item in value)
    return True



def _validate_operation(operation: dict, *, index: int) -> dict:
    if not isinstance(operation, dict):
        raise TypeError(f"Excel operation #{index} must be a dictionary.")

    _reject_macro_scope(operation, context=f"Excel operation #{index}")

    op_name_raw = operation.get("op")
    if not isinstance(op_name_raw, str) or not op_name_raw.strip():
        raise ValueError(f"Excel operation #{index} must include a non-empty 'op' string.")

    op_name = op_name_raw.strip()
    if op_name not in _SUPPORTED_OPERATIONS:
        normalized_name = op_name.lower()
        if "macro" in normalized_name or "vba" in normalized_name:
            raise ValueError(
                f"Excel operation #{index} requests unsupported VBA or macro editing scope: '{op_name}'."
            )
        raise ValueError(
            f"Excel operation #{index} has unsupported op '{op_name}'. "
            f"Supported ops: {', '.join(sorted(_SUPPORTED_OPERATIONS))}."
        )

    if op_name == "set_cell":
        sheet_name = _require_non_empty_string(operation, field="sheet", index=index)
        cell = _validate_cell_coordinate(_require_non_empty_string(operation, field="cell", index=index), index=index)
        if "value" not in operation:
            raise ValueError(f"Excel operation #{index} field 'value' is required for 'set_cell'.")
        value = operation["value"]
        validation_error = _cell_value_validation_error(value)
        if validation_error is not None:
            error_message = validation_error.args[0]
            raise type(validation_error)(
                f"Excel operation #{index} field 'value' is invalid: {error_message}"
            )
        return {"op": op_name, "sheet": sheet_name, "cell": cell, "value": value}

    if op_name == "clear_cell":
        sheet_name = _require_non_empty_string(operation, field="sheet", index=index)
        cell = _validate_cell_coordinate(_require_non_empty_string(operation, field="cell", index=index), index=index)
        return {"op": op_name, "sheet": sheet_name, "cell": cell}

    if op_name == "rename_sheet":
        from_name = _require_non_empty_string(operation, field="from", index=index)
        to_name = _require_non_empty_string(operation, field="to", index=index)
        return {"op": op_name, "from": from_name, "to": to_name}

    if op_name == "add_sheet":
        name = _require_non_empty_string(operation, field="name", index=index)
        raw_index = operation.get("index")
        if raw_index is None:
            return {"op": op_name, "name": name, "index": None}
        if not isinstance(raw_index, int):
            raise TypeError(f"Excel operation #{index} field 'index' must be an integer when provided.")
        if raw_index < 0:
            raise ValueError(f"Excel operation #{index} field 'index' must be >= 0.")
        return {"op": op_name, "name": name, "index": raw_index}

    name = _require_non_empty_string(operation, field="name", index=index)
    return {"op": op_name, "name": name}



def _require_non_empty_string(operation: dict, *, field: str, index: int) -> str:
    value = operation.get(field)
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"Excel operation #{index} field '{field}' must be a non-empty string.")
    return value.strip()



def _validate_cell_coordinate(cell: str, *, index: int) -> str:
    try:
        coordinate_to_tuple(cell)
    except Exception as exc:
        raise ValueError(f"Excel operation #{index} has invalid cell reference '{cell}'.") from exc
    return cell.upper()



def _apply_operation(workbook: Workbook, operation: dict, *, index: int) -> None:
    op_name = operation["op"]

    if op_name == "set_cell":
        sheet = _get_sheet(workbook, operation["sheet"], index=index)
        sheet[operation["cell"]] = operation["value"]
        return

    if op_name == "clear_cell":
        sheet = _get_sheet(workbook, operation["sheet"], index=index)
        sheet[operation["cell"]] = None
        return

    if op_name == "rename_sheet":
        sheet = _get_sheet(workbook, operation["from"], index=index)
        if operation["to"] in workbook.sheetnames:
            raise ValueError(
                f"Excel operation #{index} cannot rename sheet '{operation['from']}' to existing sheet '{operation['to']}'."
            )
        sheet.title = operation["to"]
        return

    if op_name == "add_sheet":
        if operation["name"] in workbook.sheetnames:
            raise ValueError(
                f"Excel operation #{index} cannot add duplicate sheet '{operation['name']}'."
            )
        add_index = operation["index"]
        if add_index is not None and add_index > len(workbook.sheetnames):
            raise ValueError(
                f"Excel operation #{index} field 'index' must be <= current sheet count {len(workbook.sheetnames)}."
            )
        workbook.create_sheet(title=operation["name"], index=add_index)
        return

    if op_name == "delete_sheet":
        sheet = _get_sheet(workbook, operation["name"], index=index)
        if len(workbook.sheetnames) <= 1:
            raise ValueError(
                f"Excel operation #{index} cannot delete sheet '{operation['name']}' because a workbook must keep at least one worksheet."
            )
        workbook.remove(sheet)
        return

    raise ValueError(f"Excel operation #{index} has unsupported op '{op_name}'.")



def _get_sheet(workbook: Workbook, sheet_name: str, *, index: int) -> Worksheet:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Excel operation #{index} references missing sheet '{sheet_name}'.")
    return workbook[sheet_name]



def _save_workbook(workbook: Workbook, *, source: Path, save_path: Path) -> None:
    source_resolved = source.resolve()
    save_resolved = save_path.resolve(strict=False)
    if source_resolved == save_resolved:
        workbook.save(save_path)
        return

    temporary_path = save_path.with_name(
        f".{save_path.stem}.tmp-{uuid4().hex}{save_path.suffix}"
    )
    try:
        workbook.save(temporary_path)
        temporary_path.replace(save_path)
    except Exception:
        if temporary_path.exists():
            temporary_path.unlink()
        raise



def _path_extension(path: Path) -> str:
    return path.suffix.lower().lstrip(".")
